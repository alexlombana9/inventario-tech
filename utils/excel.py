"""Utilidad genérica de exportación Excel con openpyxl."""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def generate_excel(
    title: str,
    headers: list[str],
    rows: list[list],
    col_widths: list[int] | None = None,
    money_cols: list[int] | None = None,
) -> BytesIO:
    """Genera un archivo Excel en memoria.

    Args:
        title: Título del reporte (fila 1).
        headers: Lista de encabezados de columna.
        rows: Lista de filas, cada una es una lista de valores.
        col_widths: Anchos de columna opcionales.
        money_cols: Índices (0-based) de columnas con formato moneda.

    Returns:
        BytesIO con el contenido del archivo .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limita a 31 chars

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell_title = ws.cell(row=1, column=1, value=title)
    cell_title.font = TITLE_FONT
    cell_title.alignment = Alignment(horizontal="center")

    # Encabezados (fila 3)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Datos
    money_cols = set(money_cols or [])
    for row_idx, row_data in enumerate(rows, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            if (col_idx - 1) in money_cols:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # Anchos de columna
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(headers[i - 1]) + 4)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
