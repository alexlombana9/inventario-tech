import os
import io
from datetime import datetime
from fastapi import APIRouter, Request, Depends, UploadFile, File, Form
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session

from database import get_db
from templates_config import templates
from auth import require_role, log_audit
import models

router = APIRouter(prefix="/importar", tags=["importar"])

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


# ── Página principal de importación ─────────────────────────────

@router.get("")
def importar_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(require_role("ADMIN")),
    msg: str = None,
    error: str = None,
):
    return templates.TemplateResponse("importar/index.html", {
        "request": request,
        "msg": msg,
        "error": error,
    })


# ── Descargar plantillas Excel ──────────────────────────────────

@router.get("/plantilla/{tipo}")
def descargar_plantilla(
    tipo: str,
    current_user: models.Usuario = Depends(require_role("ADMIN")),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    THIN_BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    plantillas = {
        "productos": {
            "headers": ["codigo", "referencia", "nombre", "descripcion", "categoria", "proveedor",
                        "precio_costo", "precio_venta", "precio_venta_minimo",
                        "stock_actual", "stock_minimo", "unidad_medida"],
            "widths": [15, 15, 30, 35, 18, 20, 16, 16, 18, 14, 14, 14],
            "ejemplo": ["PROD-001", "REF-001", "Teclado Mecanico", "Teclado gaming RGB", "Perifericos",
                        "TechDistribuidor", 45000, 89000, 75000, 25, 5, "UND"],
        },
        "clientes": {
            "headers": ["nombre", "tipo_documento", "documento", "telefono", "email", "direccion", "notas"],
            "widths": [25, 18, 18, 15, 25, 30, 30],
            "ejemplo": ["Juan Perez", "CC", "1234567890", "3001234567",
                        "juan@email.com", "Calle 123 #45-67", "Cliente frecuente"],
        },
        "proveedores": {
            "headers": ["nombre", "contacto", "telefono", "email", "direccion", "nit_ruc"],
            "widths": [25, 20, 15, 25, 30, 18],
            "ejemplo": ["TechDistribuidor S.A.", "Maria Lopez", "6011234567",
                        "ventas@techdist.com", "Av. Industrial 456", "900123456-1"],
        },
        "acreedores": {
            "headers": ["nombre", "tipo", "documento", "telefono", "email", "direccion", "notas"],
            "widths": [28, 16, 18, 15, 25, 30, 30],
            "ejemplo": ["Distribuidora XYZ", "PROVEEDOR", "900123456-1", "3001234567",
                        "cobros@xyz.com", "Av. Industrial 456", "Pago a 30 dias"],
        },
        "deudas": {
            "headers": ["concepto", "acreedor_nombre", "acreedor_tipo",
                        "monto_total", "monto_pagado", "fecha_deuda", "fecha_vencimiento", "notas"],
            "widths": [35, 25, 16, 16, 16, 16, 18, 30],
            "ejemplo": ["Compra mercancia Factura #1234", "Distribuidora XYZ", "PROVEEDOR",
                        500000, 150000, "2026-01-15", "2026-04-15", "Pago en 3 cuotas"],
        },
        "categorias": {
            "headers": ["nombre", "descripcion"],
            "widths": [30, 50],
            "ejemplo": ["Perifericos", "Teclados, mouse, audífonos y otros perifericos"],
        },
        "facturas": {
            "headers": ["numero_factura", "cliente_nombre", "cliente_documento", "cliente_telefono",
                        "cliente_email", "concepto", "monto_total", "monto_cobrado",
                        "fecha_emision", "fecha_vencimiento", "notas"],
            "widths": [16, 25, 18, 15, 22, 35, 16, 16, 16, 18, 30],
            "ejemplo": ["FAC-0001", "Juan Perez", "1234567890", "3001234567",
                        "juan@email.com", "Servicio de mantenimiento", 500000, 0,
                        "2026-01-15", "2026-02-15", "Cobro a 30 dias"],
        },
    }

    if tipo not in plantillas:
        return RedirectResponse("/importar?error=Tipo+de+plantilla+no+válido", status_code=303)

    config = plantillas[tipo]
    wb = Workbook()
    ws = wb.active
    ws.title = tipo.capitalize()

    # Encabezados
    for col_idx, header in enumerate(config["headers"], 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Fila de ejemplo
    for col_idx, val in enumerate(config["ejemplo"], 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, color="888888", italic=True)
        cell.border = THIN_BORDER

    # Anchos
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(config["widths"], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"plantilla_{tipo}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Procesar importación ────────────────────────────────────────

@router.post("/procesar")
async def procesar_importacion(
    request: Request,
    tipo: str = Form(...),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(require_role("ADMIN")),
):
    # Validar tipo de archivo
    if not archivo.filename.endswith((".xlsx", ".xls")):
        return RedirectResponse("/importar?error=Solo+se+permiten+archivos+Excel+(.xlsx,+.xls)", status_code=303)

    # Leer archivo
    content = await archivo.read()
    if len(content) > MAX_FILE_SIZE:
        return RedirectResponse("/importar?error=El+archivo+excede+el+tamaño+máximo+(10MB)", status_code=303)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return RedirectResponse("/importar?error=No+se+pudo+leer+el+archivo+Excel", status_code=303)

    # Leer encabezados
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return RedirectResponse("/importar?error=El+archivo+está+vacío+o+solo+tiene+encabezados", status_code=303)

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    # Despachar según tipo
    if tipo == "productos":
        result = _importar_productos(db, headers, data_rows, current_user, request)
    elif tipo == "clientes":
        result = _importar_clientes(db, headers, data_rows, current_user, request)
    elif tipo == "proveedores":
        result = _importar_proveedores(db, headers, data_rows, current_user, request)
    elif tipo == "acreedores":
        result = _importar_acreedores(db, headers, data_rows, current_user, request)
    elif tipo == "deudas":
        result = _importar_deudas(db, headers, data_rows, current_user, request)
    elif tipo == "categorias":
        result = _importar_categorias(db, headers, data_rows, current_user, request)
    elif tipo == "facturas":
        result = _importar_facturas(db, headers, data_rows, current_user, request)
    else:
        return RedirectResponse("/importar?error=Tipo+de+importación+no+válido", status_code=303)

    wb.close()
    return result


def _col_index(headers: list, name: str) -> int | None:
    """Busca el índice de una columna por nombre."""
    try:
        return headers.index(name)
    except ValueError:
        return None


def _cell_str(row, idx) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


def _cell_float(row, idx, default=0.0) -> float:
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


# ── Importar Productos ──────────────────────────────────────────

def _importar_productos(db: Session, headers: list, data_rows: list, user, request: Request):
    col_codigo = _col_index(headers, "codigo")
    col_ref = _col_index(headers, "referencia")
    col_nombre = _col_index(headers, "nombre")
    col_desc = _col_index(headers, "descripcion")
    col_cat = _col_index(headers, "categoria")
    col_prov = _col_index(headers, "proveedor")
    col_pcosto = _col_index(headers, "precio_costo")
    col_pventa = _col_index(headers, "precio_venta")
    col_pvmin = _col_index(headers, "precio_venta_minimo")
    col_stock = _col_index(headers, "stock_actual")
    col_smin = _col_index(headers, "stock_minimo")
    col_um = _col_index(headers, "unidad_medida")

    if col_codigo is None or col_nombre is None:
        return RedirectResponse(
            "/importar?error=El+archivo+debe+tener+las+columnas+'codigo'+y+'nombre'",
            status_code=303,
        )

    creados = 0
    actualizados = 0
    errores = []
    cat_cache = {}
    prov_cache = {}

    for i, row in enumerate(data_rows, start=2):
        codigo = _cell_str(row, col_codigo)
        nombre = _cell_str(row, col_nombre)

        if not codigo or not nombre:
            errores.append(f"Fila {i}: código o nombre vacío, omitida")
            continue

        # Buscar o crear categoría
        categoria_id = None
        cat_nombre = _cell_str(row, col_cat)
        if cat_nombre:
            if cat_nombre not in cat_cache:
                cat_obj = db.query(models.Categoria).filter(
                    models.Categoria.nombre.ilike(cat_nombre)
                ).first()
                if not cat_obj:
                    cat_obj = models.Categoria(nombre=cat_nombre)
                    db.add(cat_obj)
                    db.flush()
                cat_cache[cat_nombre] = cat_obj.id
            categoria_id = cat_cache[cat_nombre]

        # Buscar proveedor
        proveedor_id = None
        prov_nombre = _cell_str(row, col_prov)
        if prov_nombre:
            if prov_nombre not in prov_cache:
                prov_obj = db.query(models.Proveedor).filter(
                    models.Proveedor.nombre.ilike(prov_nombre)
                ).first()
                prov_cache[prov_nombre] = prov_obj.id if prov_obj else None
            proveedor_id = prov_cache[prov_nombre]

        # Buscar producto existente por código
        existente = db.query(models.Producto).filter(models.Producto.codigo == codigo).first()

        if existente:
            existente.nombre = nombre
            ref = _cell_str(row, col_ref)
            if ref:
                existente.referencia = ref
            existente.descripcion = _cell_str(row, col_desc) or existente.descripcion
            if categoria_id:
                existente.categoria_id = categoria_id
            if proveedor_id:
                existente.proveedor_id = proveedor_id
            pc = _cell_float(row, col_pcosto)
            pv = _cell_float(row, col_pventa)
            pvmin = _cell_float(row, col_pvmin)
            if pc > 0:
                existente.precio_costo = pc
            if pv > 0:
                existente.precio_venta = pv
            if pvmin > 0:
                existente.precio_venta_minimo = pvmin
            stock_new = _cell_float(row, col_stock)
            if col_stock is not None and stock_new > 0:
                existente.stock_actual = stock_new
            smin = _cell_float(row, col_smin)
            if smin > 0:
                existente.stock_minimo = smin
            um = _cell_str(row, col_um)
            if um:
                existente.unidad_medida = um
            actualizados += 1
        else:
            stock_val = _cell_float(row, col_stock)
            producto = models.Producto(
                codigo=codigo,
                referencia=_cell_str(row, col_ref),
                nombre=nombre,
                descripcion=_cell_str(row, col_desc),
                categoria_id=categoria_id,
                proveedor_id=proveedor_id,
                precio_costo=_cell_float(row, col_pcosto),
                precio_venta=_cell_float(row, col_pventa),
                precio_venta_minimo=_cell_float(row, col_pvmin),
                stock_actual=stock_val,
                stock_minimo=_cell_float(row, col_smin),
                unidad_medida=_cell_str(row, col_um) or "UND",
            )
            db.add(producto)
            db.flush()

            # Registrar movimiento de entrada inicial si tiene stock
            if stock_val > 0:
                mov = models.MovimientoInventario(
                    producto_id=producto.id,
                    tipo="ENTRADA",
                    cantidad=stock_val,
                    stock_anterior=0,
                    stock_resultante=stock_val,
                    precio_unitario=_cell_float(row, col_pcosto),
                    observaciones="Importación desde Excel",
                )
                db.add(mov)

            creados += 1

    db.commit()

    ip = request.client.host if request.client else ""
    log_audit(db, user, "CREATE", "importacion", None,
              f"Importación productos: {creados} creados, {actualizados} actualizados", ip)

    msg_parts = []
    if creados:
        msg_parts.append(f"{creados} productos creados")
    if actualizados:
        msg_parts.append(f"{actualizados} actualizados")
    if errores:
        msg_parts.append(f"{len(errores)} errores")

    msg = "Importación completada: " + ", ".join(msg_parts) if msg_parts else "No se importaron datos"
    return RedirectResponse(f"/importar?msg={msg.replace(' ', '+')}", status_code=303)


# ── Importar Clientes ───────────────────────────────────────────

def _importar_clientes(db: Session, headers: list, data_rows: list, user, request: Request):
    col_nombre = _col_index(headers, "nombre")
    col_tipodoc = _col_index(headers, "tipo_documento")
    col_doc = _col_index(headers, "documento")
    col_tel = _col_index(headers, "telefono")
    col_email = _col_index(headers, "email")
    col_dir = _col_index(headers, "direccion")
    col_notas = _col_index(headers, "notas")

    if col_nombre is None:
        return RedirectResponse(
            "/importar?error=El+archivo+debe+tener+la+columna+'nombre'",
            status_code=303,
        )

    creados = 0
    omitidos = 0

    for i, row in enumerate(data_rows, start=2):
        nombre = _cell_str(row, col_nombre)
        if not nombre:
            omitidos += 1
            continue

        documento = _cell_str(row, col_doc)

        # Verificar duplicado por documento si existe
        if documento:
            existente = db.query(models.Cliente).filter(
                models.Cliente.documento == documento
            ).first()
            if existente:
                omitidos += 1
                continue

        cliente = models.Cliente(
            nombre=nombre,
            tipo_documento=_cell_str(row, col_tipodoc) or "CC",
            documento=documento,
            telefono=_cell_str(row, col_tel),
            email=_cell_str(row, col_email),
            direccion=_cell_str(row, col_dir),
            notas=_cell_str(row, col_notas),
        )
        db.add(cliente)
        creados += 1

    db.commit()

    ip = request.client.host if request.client else ""
    log_audit(db, user, "CREATE", "importacion", None,
              f"Importación clientes: {creados} creados, {omitidos} omitidos", ip)

    msg = f"Importación completada: {creados} clientes creados"
    if omitidos:
        msg += f", {omitidos} omitidos (duplicados o vacíos)"
    return RedirectResponse(f"/importar?msg={msg.replace(' ', '+')}", status_code=303)


# ── Importar Proveedores ────────────────────────────────────────

def _importar_proveedores(db: Session, headers: list, data_rows: list, user, request: Request):
    col_nombre = _col_index(headers, "nombre")
    col_contacto = _col_index(headers, "contacto")
    col_tel = _col_index(headers, "telefono")
    col_email = _col_index(headers, "email")
    col_dir = _col_index(headers, "direccion")
    col_nit = _col_index(headers, "nit_ruc")

    if col_nombre is None:
        return RedirectResponse(
            "/importar?error=El+archivo+debe+tener+la+columna+'nombre'",
            status_code=303,
        )

    creados = 0
    omitidos = 0

    for i, row in enumerate(data_rows, start=2):
        nombre = _cell_str(row, col_nombre)
        if not nombre:
            omitidos += 1
            continue

        # Verificar duplicado por nombre
        existente = db.query(models.Proveedor).filter(
            models.Proveedor.nombre.ilike(nombre)
        ).first()
        if existente:
            omitidos += 1
            continue

        proveedor = models.Proveedor(
            nombre=nombre,
            contacto=_cell_str(row, col_contacto),
            telefono=_cell_str(row, col_tel),
            email=_cell_str(row, col_email),
            direccion=_cell_str(row, col_dir),
            nit_ruc=_cell_str(row, col_nit),
        )
        db.add(proveedor)
        creados += 1

    db.commit()

    ip = request.client.host if request.client else ""
    log_audit(db, user, "CREATE", "importacion", None,
              f"Importación proveedores: {creados} creados, {omitidos} omitidos", ip)

    msg = f"Importación completada: {creados} proveedores creados"
    if omitidos:
        msg += f", {omitidos} omitidos (duplicados o vacíos)"
    return RedirectResponse(f"/importar?msg={msg.replace(' ', '+')}", status_code=303)


# ── Importar Acreedores ────────────────────────────────────────

def _importar_acreedores(db: Session, headers: list, data_rows: list, user, request: Request):
    col_nombre = _col_index(headers, "nombre")
    col_tipo = _col_index(headers, "tipo")
    col_doc = _col_index(headers, "documento")
    col_tel = _col_index(headers, "telefono")
    col_email = _col_index(headers, "email")
    col_dir = _col_index(headers, "direccion")
    col_notas = _col_index(headers, "notas")

    if col_nombre is None:
        return RedirectResponse(
            "/importar?error=El+archivo+debe+tener+la+columna+'nombre'",
            status_code=303,
        )

    TIPOS_VALIDOS = {"PROVEEDOR", "BANCO", "PERSONA", "OTRO"}
    creados = 0
    omitidos = 0

    for i, row in enumerate(data_rows, start=2):
        nombre = _cell_str(row, col_nombre)
        if not nombre:
            omitidos += 1
            continue

        # Verificar duplicado por nombre
        existente = db.query(models.Acreedor).filter(
            models.Acreedor.nombre.ilike(nombre),
            models.Acreedor.activo == True,
        ).first()
        if existente:
            omitidos += 1
            continue

        tipo = _cell_str(row, col_tipo).upper()
        if tipo not in TIPOS_VALIDOS:
            tipo = "OTRO"

        acreedor = models.Acreedor(
            nombre=nombre,
            tipo=tipo,
            documento=_cell_str(row, col_doc),
            telefono=_cell_str(row, col_tel),
            email=_cell_str(row, col_email),
            direccion=_cell_str(row, col_dir),
            notas=_cell_str(row, col_notas),
        )
        db.add(acreedor)
        creados += 1

    db.commit()

    ip = request.client.host if request.client else ""
    log_audit(db, user, "CREATE", "importacion", None,
              f"Importación acreedores: {creados} creados, {omitidos} omitidos", ip)

    msg = f"Importación completada: {creados} acreedores creados"
    if omitidos:
        msg += f", {omitidos} omitidos (duplicados o vacíos)"
    return RedirectResponse(f"/importar?msg={msg.replace(' ', '+')}", status_code=303)


# ── Importar Deudas ────────────────────────────────────────────

def _importar_deudas(db: Session, headers: list, data_rows: list, user, request: Request):
    col_concepto = _col_index(headers, "concepto")
    col_acr_nombre = _col_index(headers, "acreedor_nombre")
    col_acr_tipo = _col_index(headers, "acreedor_tipo")
    col_monto = _col_index(headers, "monto_total")
    col_pagado = _col_index(headers, "monto_pagado")
    col_fecha = _col_index(headers, "fecha_deuda")
    col_venc = _col_index(headers, "fecha_vencimiento")
    col_notas = _col_index(headers, "notas")

    if col_concepto is None or col_acr_nombre is None or col_monto is None:
        return RedirectResponse(
            "/importar?error=El+archivo+debe+tener+las+columnas+'concepto',+'acreedor_nombre'+y+'monto_total'",
            status_code=303,
        )

    TIPOS_VALIDOS = {"PROVEEDOR", "BANCO", "PERSONA", "OTRO"}
    creados = 0
    errores = 0
    acr_cache = {}

    for i, row in enumerate(data_rows, start=2):
        concepto = _cell_str(row, col_concepto)
        acr_nombre = _cell_str(row, col_acr_nombre)
        monto_total = _cell_float(row, col_monto)

        if not concepto or not acr_nombre or monto_total <= 0:
            errores += 1
            continue

        acr_tipo = _cell_str(row, col_acr_tipo).upper()
        if acr_tipo not in TIPOS_VALIDOS:
            acr_tipo = "OTRO"

        # Buscar acreedor registrado por nombre (cache)
        acreedor_id = None
        if acr_nombre not in acr_cache:
            acr_obj = db.query(models.Acreedor).filter(
                models.Acreedor.nombre.ilike(acr_nombre),
                models.Acreedor.activo == True,
            ).first()
            acr_cache[acr_nombre] = acr_obj.id if acr_obj else None
        acreedor_id = acr_cache[acr_nombre]

        # Parsear fechas
        fecha_deuda = _parse_date(_cell_str(row, col_fecha))
        fecha_venc = _parse_date(_cell_str(row, col_venc))

        monto_pagado = _cell_float(row, col_pagado)
        if monto_pagado > monto_total:
            monto_pagado = monto_total

        # Calcular estado
        if monto_pagado >= monto_total:
            estado = "PAGADO"
        elif monto_pagado > 0:
            estado = "PARCIAL"
        else:
            estado = "PENDIENTE"

        deuda = models.Deuda(
            concepto=concepto,
            acreedor_nombre=acr_nombre,
            acreedor_tipo=acr_tipo,
            acreedor_id=acreedor_id,
            monto_total=monto_total,
            monto_pagado=monto_pagado,
            fecha_deuda=fecha_deuda or datetime.now(),
            fecha_vencimiento=fecha_venc,
            estado=estado,
            notas=_cell_str(row, col_notas),
        )
        db.add(deuda)
        creados += 1

    db.commit()

    ip = request.client.host if request.client else ""
    log_audit(db, user, "CREATE", "importacion", None,
              f"Importación deudas: {creados} creadas, {errores} errores", ip)

    msg_parts = []
    if creados:
        msg_parts.append(f"{creados} deudas creadas")
    if errores:
        msg_parts.append(f"{errores} filas omitidas")
    msg = "Importación completada: " + ", ".join(msg_parts) if msg_parts else "No se importaron datos"
    return RedirectResponse(f"/importar?msg={msg.replace(' ', '+')}", status_code=303)


# ── Importar Categorías ──────────────────────────────────────

def _importar_categorias(db: Session, headers: list, data_rows: list, user, request: Request):
    col_nombre = _col_index(headers, "nombre")
    col_desc = _col_index(headers, "descripcion")

    if col_nombre is None:
        return RedirectResponse(
            "/importar?error=El+archivo+debe+tener+la+columna+'nombre'",
            status_code=303,
        )

    creados = 0
    omitidos = 0

    for i, row in enumerate(data_rows, start=2):
        nombre = _cell_str(row, col_nombre)
        if not nombre:
            omitidos += 1
            continue

        existente = db.query(models.Categoria).filter(
            models.Categoria.nombre.ilike(nombre)
        ).first()
        if existente:
            omitidos += 1
            continue

        categoria = models.Categoria(
            nombre=nombre,
            descripcion=_cell_str(row, col_desc),
        )
        db.add(categoria)
        creados += 1

    db.commit()

    ip = request.client.host if request.client else ""
    log_audit(db, user, "CREATE", "importacion", None,
              f"Importacion categorias: {creados} creadas, {omitidos} omitidas", ip)

    msg = f"Importacion completada: {creados} categorias creadas"
    if omitidos:
        msg += f", {omitidos} omitidas (duplicadas o vacias)"
    return RedirectResponse(f"/importar?msg={msg.replace(' ', '+')}", status_code=303)


# ── Importar Facturas / Cuentas por Cobrar ──────────────────

def _importar_facturas(db: Session, headers: list, data_rows: list, user, request: Request):
    col_num = _col_index(headers, "numero_factura")
    col_cliente = _col_index(headers, "cliente_nombre")
    col_doc = _col_index(headers, "cliente_documento")
    col_tel = _col_index(headers, "cliente_telefono")
    col_email = _col_index(headers, "cliente_email")
    col_concepto = _col_index(headers, "concepto")
    col_monto = _col_index(headers, "monto_total")
    col_cobrado = _col_index(headers, "monto_cobrado")
    col_emision = _col_index(headers, "fecha_emision")
    col_venc = _col_index(headers, "fecha_vencimiento")
    col_notas = _col_index(headers, "notas")

    if col_cliente is None or col_concepto is None or col_monto is None:
        return RedirectResponse(
            "/importar?error=El+archivo+debe+tener+las+columnas+'cliente_nombre',+'concepto'+y+'monto_total'",
            status_code=303,
        )

    creados = 0
    omitidos = 0
    errores = 0

    # Generar siguiente numero si no se proporciona columna
    ultimo = db.query(models.Factura).order_by(models.Factura.id.desc()).first()
    next_num = 1
    if ultimo:
        try:
            next_num = int(ultimo.numero_factura.split("-")[-1]) + 1
        except (ValueError, IndexError):
            next_num = ultimo.id + 1

    for i, row in enumerate(data_rows, start=2):
        cliente = _cell_str(row, col_cliente)
        concepto = _cell_str(row, col_concepto)
        monto_total = _cell_float(row, col_monto)

        if not cliente or not concepto or monto_total <= 0:
            errores += 1
            continue

        # Numero de factura
        num_factura = _cell_str(row, col_num)
        if not num_factura:
            num_factura = f"FAC-{next_num:04d}"
            next_num += 1

        # Verificar duplicado por numero_factura
        existente = db.query(models.Factura).filter(
            models.Factura.numero_factura == num_factura
        ).first()
        if existente:
            omitidos += 1
            continue

        # Parsear fechas
        fecha_emision = _parse_date(_cell_str(row, col_emision))
        fecha_venc = _parse_date(_cell_str(row, col_venc))

        monto_cobrado = _cell_float(row, col_cobrado)
        if monto_cobrado > monto_total:
            monto_cobrado = monto_total

        # Calcular estado
        if monto_cobrado >= monto_total:
            estado = "PAGADO"
        elif monto_cobrado > 0:
            estado = "PARCIAL"
        else:
            estado = "PENDIENTE"

        factura = models.Factura(
            numero_factura=num_factura,
            cliente_nombre=cliente,
            cliente_documento=_cell_str(row, col_doc),
            cliente_telefono=_cell_str(row, col_tel),
            cliente_email=_cell_str(row, col_email),
            concepto=concepto,
            monto_total=monto_total,
            monto_cobrado=monto_cobrado,
            fecha_emision=fecha_emision or datetime.now(),
            fecha_vencimiento=fecha_venc,
            estado=estado,
            notas=_cell_str(row, col_notas),
        )
        db.add(factura)
        creados += 1
        next_num += 1

    db.commit()

    ip = request.client.host if request.client else ""
    log_audit(db, user, "CREATE", "importacion", None,
              f"Importación facturas: {creados} creadas, {omitidos} omitidas, {errores} errores", ip)

    msg_parts = []
    if creados:
        msg_parts.append(f"{creados} facturas creadas")
    if omitidos:
        msg_parts.append(f"{omitidos} omitidas (duplicadas)")
    if errores:
        msg_parts.append(f"{errores} filas con datos incompletos")
    msg = "Importación completada: " + ", ".join(msg_parts) if msg_parts else "No se importaron datos"
    return RedirectResponse(f"/importar?msg={msg.replace(' ', '+')}", status_code=303)


def _parse_date(val: str):
    """Intenta parsear una fecha en varios formatos comunes."""
    if not val:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(val.strip(), fmt)
        except ValueError:
            continue
    return None
